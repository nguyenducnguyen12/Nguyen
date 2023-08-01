##quản lý thư viện v1.2
#modules
import customtkinter as ctk
import tkinter,webbrowser
from tkinter import *
from tkinter import messagebox,ttk
from datetime import date
import os,pathlib,openpyxl,xlrd
from openpyxl import Workbook
import requests,sys
from bs4 import BeautifulSoup
import openpyxl
from pathlib import Path
from PIL import Image
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")
ds_admin=[]
my_file = Path("admin.txt")
if not my_file.is_file():
    my_file.touch()
with open('admin.txt','w') as n:
    n.write('Admin')
    n.close()
with open('admin.txt','r') as f:
    f=f.read()
    a=f.split(',')
    for i in range(len(a)):
        ds_admin.append(a[i])
file=pathlib.Path('DANH_SACH_MUON.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']='Tên'
    sheet['B1']='Lớp'
    sheet['C1']='Thông Tin Liên Lạc'
    sheet['D1']='Tên Sách Mươn'
    sheet['E1']='Ngày Mượn:'
    sheet['F1']='Hạn'
    sheet['G1']='Người Xét Duyệt'
    file.save('DANH_SACH_MUON.xlsx')
class CTKAPP(ctk.CTk):
	def __init__(self, *args, **kwargs):
		ctk.CTk.__init__(self, *args, **kwargs)
		container = ctk.CTkFrame(self)
		container.pack(side = "top", fill = "both", expand = True)
		container.grid_rowconfigure(0, weight = 1)
		container.grid_columnconfigure(0, weight = 1)
		self.frames = {}
		for F in (StartPage, Page1, Page2,ADMIN,Duyet):
			frame = F(container, self)
			self.frames[F] = frame
			frame.grid(row = 0, column = 0, sticky ="nsew")
		self.show_frame(StartPage)
	def show_frame(self, cont):
		frame = self.frames[cont]
		frame.tkraise()
class StartPage(ctk.CTkFrame):
	def __init__(self, parent, controller):
		ctk.CTkFrame.__init__(self, parent)
		frame1=ctk.CTkFrame(self,width=587,height=30)
		frame1.place(x=110,y=472)
		bt1=ctk.CTkButton(frame1,text='Home',command=lambda:controller.show_frame(StartPage))
		bt1.place(x=0,y=0)
		
		bt=ctk.CTkButton(frame1,text='Duyệt Người Mượn',command=lambda:controller.show_frame(Duyet))
		bt.place(x=145,y=0)
		bt2=ctk.CTkButton(frame1,text='Danh sách Người Mượn',command=lambda:controller.show_frame(Page1))
		bt2.place(x=290,y=0)
		bt3=ctk.CTkButton(frame1,text='Cài Đặt',command=lambda:controller.show_frame(Page2))
		bt3.place(x=445,y=0)
		b=ctk.CTkLabel(self)
		b.place(x=60,y=30)
		image=ctk.CTkImage(light_image=Image.open("loa.png"),
                                  size=(30, 30))
		ctk.CTkLabel(self,image=image,text='').place(x=30,y=30)
		frame2=ctk.CTkFrame(self,width=200,height=100)
		frame2.pack_configure(side='right')
		a=ctk.CTkLabel(frame2,text='Thông Báo')
		a.place(x=40,y=0)
		thongbao1=ctk.CTkLabel(frame2,text='Chưa có thông báo mới')
		thongbao1.place(x=10,y=20)
		gach=ctk.CTkLabel(frame2,text='--------------------------------')
		gach.place(x=40,y=50)
		thongbao2=ctk.CTkLabel(frame2,text='chưa có bạn nào đến hạn'
			 )
		thongbao2.place(x=10,y=70)
		def MovingText(s):
			s1 = s[1:len(s)]
			s2 = s[0:1]
			string = s1 + s2
			b.configure(text = string)
			b.after(500,MovingText,string)
	
		ca='         Bản cập nhật 1.2 mới nhất đã cập nhật thành công'
		MovingText(ca)
		
class Page1(ctk.CTkFrame):
	def __init__(self, parent, controller):
		ctk.CTkFrame.__init__(self, parent)
		
		treeview =ttk.Treeview(self, columns=( 'ten', 'lop', 'tensach','ngaymuon','han'))
		treeview.column("#0", width=5)
		treeview.column('ten', anchor="w", width=100)
		treeview.column('lop', anchor="w", width=5)
		treeview.column("tensach", width=100)
		treeview.column("ngaymuon", width=100)
		treeview.column('han', anchor="w", width=100)

		treeview.heading("#0",text='Số thứ tự',anchor='center')
		treeview.heading('ten', text='Tên')
		treeview.heading('lop', text='Lớp')
		treeview.heading('tensach', text='Tên sách')
		treeview.heading('ngaymuon',text='Ngày Mượn')
		treeview.heading('han', text='Hạn')
		treeview.pack(side='left', fill='both', expand=True)
		style = ttk.Style()
		style.theme_use('clam')
		style.configure('Treeview', background='#333333', foreground='white', fieldbackground='#333333', rowheight=25)
		style.configure('Treeview.Heading', background='#333333', foreground='white')
		# Create the vertical scrollbar
		scrollbar = ctk.CTkScrollbar(self, command=treeview.yview)
		scrollbar.pack(side='right', fill='y')
		treeview.configure(yscrollcommand=scrollbar.set)

		# Load data from Excel file
		wb = openpyxl.load_workbook('DANH_SACH_MUON.xlsx')
		ws = wb.active
		for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
			stt = i + 1
			ten = row[0]
			lop = row[1]
			tensach=row[3]
			ngaymuon=row[4]
			han = row[5]
			treeview.insert('', 'end', text=str(stt), values=(ten, lop, tensach,ngaymuon,han))
		def on_select(event):
			selected_item = treeview.focus()
			if selected_item:
				msg = f'Bạn có chắc muốn xóa {treeview.item(selected_item, "values")}?'
				confirm = messagebox.askyesno('Xác nhận', msg)
				if confirm:
                    # Remove the item from the Treeview
					treeview.delete(selected_item)

					# Update the data in the Excel file
					row_index = int(selected_item[1:]) - 1
					ws.delete_rows(row_index + 2)
					wb.save('DANH_SACH_MUON.xlsx')
		treeview.bind('<<TreeviewSelect>>', on_select)
		frame1=ctk.CTkFrame(self,width=587,height=30)
		frame1.place(x=110,y=472)
		bt1=ctk.CTkButton(frame1,text='Home',command=lambda:controller.show_frame(StartPage))
		bt1.place(x=0,y=0)
		bt=ctk.CTkButton(frame1,text='Duyệt Người Mượn',command=lambda:controller.show_frame(Duyet))
		bt.place(x=145,y=0)
		bt2=ctk.CTkButton(frame1,text='Danh sách Người Mượn',command=lambda:messagebox.showinfo('Information','Bạn Đang ở textnorans'))
		bt2.place(x=290,y=0)
		bt3=ctk.CTkButton(frame1,text='Cài Đặt',command=lambda:controller.show_frame(Page2))
		bt3.place(x=445,y=0)
class Page2(ctk.CTkFrame):
	def __init__(self, parent, controller):
		ctk.CTkFrame.__init__(self, parent)
		ctk.CTkLabel(self,text='Cài Đặt').pack_configure(side='top')
		
		frame2=ctk.CTkFrame(self)
		frame2.pack_configure(side='left')
		frame1=ctk.CTkFrame(self,width=587,height=30)
		frame1.place(x=110,y=472)
		bt1=ctk.CTkButton(frame1,text='Home',command=lambda:controller.show_frame(StartPage))
		bt1.place(x=0,y=0)
		bt=ctk.CTkButton(frame1,text='Duyệt Người Mượn',command=lambda:controller.show_frame(Duyet))
		bt.place(x=145,y=0)
		bt2=ctk.CTkButton(frame1,text='Danh sách Người Mượn',command=lambda:controller.show_frame(Page1))
		bt2.place(x=290,y=0)
		bt3=ctk.CTkButton(frame1,text='Cài Đặt',command=lambda:messagebox.showinfo('Information','Bạn Đang ở setting'))
		bt3.place(x=445,y=0)
		def reset_all():
			msg=messagebox.askquestion('Infomation','Khôi phục cài đặt gốc sẽ xoá toàn bộ dữ liệu kể cả dữ liệu    người mượn.Xác nhận xoá?')
			if msg=='yes':
				os.remove('DANH_SACH_MUON.xlsx')
				os.remove('admin.txt')
				messagebox.showinfo('complete','Đã Khôi Phục Thành Công Cài Đặt Gốc')
				app.destroy()
			elif msg=='no':
				messagebox.showinfo('Cancel','Đã Huỷ khôi phục')
		frame_bt1=ctk.CTkButton(frame2,text='Khôi Phục Cài Đặt Gốc',command=reset_all)
		frame_bt1.pack()
		frame_bt2=ctk.CTkButton(frame2,text='Quản lý Admin',command=lambda:controller.show_frame(ADMIN))
		frame_bt2.pack(pady=15)
		frame_bt3=ctk.CTkButton(frame2,text='Kiểm tra cập nhật')
		frame_bt3.pack(pady=10)
class ADMIN(ctk.CTkFrame):
	def __init__(self, parent, controller):
		ctk.CTkFrame.__init__(self, parent)
		lb1=ctk.CTkLabel(self,text='Quản Lý Admin')
		lb1.pack()
		frame1=ctk.CTkFrame(self,width=587,height=30)
		frame1.place(x=110,y=472)
		bt1=ctk.CTkButton(frame1,text='Home',command=lambda:messagebox.showinfo('Information','Bạn Đang ở Home'))
		bt1.place(x=0,y=0)
		bt=ctk.CTkButton(frame1,text='Duyệt Người Mượn',command=lambda:controller.show_frame(Duyet))
		bt.place(x=145,y=0)
		bt2=ctk.CTkButton(frame1,text='Danh sách Người Mượn',command=lambda:controller.show_frame(Page1))
		bt2.place(x=290,y=0)
		bt3=ctk.CTkButton(frame1,text='Cài Đặt',command=lambda:controller.show_frame(Page2))
		bt3.place(x=445,y=0)
		#tree
		tree=ttk.Treeview(self,columns=('ten','maadmin','xacthuc'))
		tree.column("#0",width=2)
		tree.column("ten",anchor='w',width=100)
		tree.column('maadmin',anchor='w',width=10)
		tree.column('xacthuc',width=60)
		tree.heading("#0",text='Số thứ tự',anchor='center')
		tree.heading('ten', text='Tên')
		tree.heading('maadmin', text='Mã Xác Thực')
		tree.heading('xacthuc', text='Mã ADMIN')
		tree.place(x=40,y=10)
		style = ttk.Style()
		style.theme_use('clam')
		style.configure('Treeview', background='#333333', foreground='white', fieldbackground='#333333', rowheight=25)
		style.configure('Treeview.Heading', background='#333333', foreground='white')
		# Create the vertical scrollbar
		scrollbar = ctk.CTkScrollbar(self, command=tree.yview)
		scrollbar.pack(side='right', fill='y')
		tree.configure(yscrollcommand=scrollbar.set)
class Duyet(ctk.CTkFrame):
	
	def __init__(self, parent, controller):
		colortext='white'
		today=date.today()
		d1=today.strftime("%d/%m/20%y")
		ctk.CTkFrame.__init__(self, parent)
		from tkcalendar import DateEntry
		ten_lb=ctk.CTkLabel(self,text='Tên',text_color=colortext)
		ten_lb.place(x=50,y=30)
		lop_lb=ctk.CTkLabel(self,text='Lớp',text_color=colortext)
		lop_lb.place(x=50,y=90)
		cto_lb=ctk.CTkLabel(self,text='Thông Tin Liên Lạc',text_color=colortext)
		cto_lb.place(x=50,y=150)
		book_name_lb=ctk.CTkLabel(self,text='Tên Sách',text_color=colortext)
		book_name_lb.place(x=470,y=30)
		han_lb=ctk.CTkLabel(self,text='Hạn',text_color=colortext)
		han_lb.place(x=470,y=90)
		admin_duyet_lb=ctk.CTkLabel(self,text='Người Duyệt',text_color=colortext)
		admin_duyet_lb.place(x=470,y=150)
		#entry
		name_et=ctk.CTkEntry(self,placeholder_text='Tên')
		name_et.place(x=200,y=30)
		lop_et=ctk.CTkEntry(self,placeholder_text='Lớp')
		lop_et.place(x=200,y=90)
		contact_et=ctk.CTkEntry(self,placeholder_text='CONTACT')
		contact_et.place(x=200,y=150)
		book_name_et=ctk.CTkEntry(self,placeholder_text='Tên Sách')
		book_name_et.place(x=640,y=30)

		han_et = DateEntry(self, width=12, background='darkblue',
                                foreground='white', borderwidth=2)
		han_et.place(x=640, y=90)
		admin_op=ctk.CTkOptionMenu(self,values=ds_admin)
		admin_op.place(x=640,y=150)
		admin_op.set('Chọn Người Xét')
		ctk.CTkButton(self, text='Lưu Thông tin', command=lambda: save_data()).place(x=300, y=300)
		def save_data():
			wb = openpyxl.load_workbook('DANH_SACH_MUON.xlsx')
			sheet = wb.active
			name = name_et.get()
			lop = lop_et.get()
			contact = contact_et.get()
			book_name = book_name_et.get()
			han = han_et.get()
			nguoiduyet = admin_op.get()
			if name=='':
				messagebox.showerror("Lỗi","Bạn chưa nhập tên")
			if lop=='':
				messagebox.showerror("Lỗi", "Bạn chưa nhập lớp")
			elif contact=='':
				messagebox.showerror("Lỗi","Bạn chưa nhập thông tin liên lạc")
			elif book_name=='':
				messagebox.showerror("Lỗi", "Bạn chưa nhập tên sách")
			elif han=='':
				messagebox.showerror("Lỗi", "Bạn chưa điền hạn trả sách")
			elif nguoiduyet=='':
				messagebox.showerror("Lỗi","Bạn chưa chọn người duyệt")
			else:
				row_num = sheet.max_row + 1
			# Write the data to the appropriate cells
				sheet.cell(row=row_num, column=1, value=name)
				sheet.cell(row=row_num, column=2, value=lop)
				sheet.cell(row=row_num, column=3, value=contact)
				sheet.cell(row=row_num, column=4, value=book_name)
				sheet.cell(row=row_num, column=5, value=d1)
				

				sheet.cell(row=row_num, column=6, value=han)
				sheet.cell(row=row_num, column=7, value=nguoiduyet)
				messagebox.showinfo('Thành Công','Đã Lưu Thành Công')
		        # Save the Excel file
				wb.save('DANH_SACH_MUON.xlsx')
				app.destroy()
				os.system('py version2.py')
		frame1=ctk.CTkFrame(self,width=587,height=30)
		frame1.place(x=110,y=472)
		bt1=ctk.CTkButton(frame1,text='Home',command=lambda:controller.show_frame(StartPage))
		bt1.place(x=0,y=0)
		bt=ctk.CTkButton(frame1,text='Duyệt Người Mượn',command=lambda:controller.show_frame(Duyet))
		bt.place(x=145,y=0)
		bt2=ctk.CTkButton(frame1,text='Danh sách Người Mượn',command=lambda:controller.show_frame(Page1))
		bt2.place(x=290,y=0)
		bt3=ctk.CTkButton(frame1,text='Cài Đặt',command=lambda:controller.show_frame(Page2))
		bt3.place(x=445,y=0)
app = CTKAPP()
app.title('Quản lý thư viện')
app.iconbitmap('logo.ico')
app.resizable(False,False)
app.geometry('800x500')
app.mainloop()
